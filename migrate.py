"""
Existem dois bancos de dados (banco1.csv e banco2.csv)

O banco1 está desatualizado e necessita ser atualizado baseado no banco2.

O problema é que o banco1 e o banco2 não tem chaves primárias em comum, a única forma de identificar um item equivalente é pelo nome.

Os dados no banco2 foram inseridos manualmente e apresentam erros de ortografia nos produtos.

Crie um terceiro banco com as descrições do banco1 e valores do banco2.

Linguagem: Python 3.6.1
Objetivo: Entregar um código capaz de resolver o problema acima escrevendo os dados para um arquivo banco3.
"""
import abc
import argparse
import functools
import unicodedata

from decimal import Decimal
from typing import NamedTuple, Generator, Any, Iterable

from fuzzywuzzy import fuzz
from openpyxl import load_workbook, Workbook


MIN_RATIO_EQUIVALENCE = 90


def get_normalize_description(description) -> str:
    normalized = unicodedata.normalize('NFKD', description)
    return normalized.encode('ascii', 'ignore')


@functools.lru_cache(maxsize=50)
def calc_description_equivalence_ratio(desc1, desc2) -> float:
    ratio = fuzz.token_sort_ratio(
        get_normalize_description(desc1),
        get_normalize_description(desc2)
    )
    return ratio


class Product(NamedTuple):
    description: str
    price: Decimal

    def __eq__(self, product) -> bool:
        ratio = calc_description_equivalence_ratio(
            self.description.lower(),
            product.description.lower()
        )
        return ratio >= MIN_RATIO_EQUIVALENCE


class BaseReadDB(metaclass=abc.ABCMeta):

    def __init__(self, filename):
        self.filename = filename

    @abc.abstractmethod
    def read(self) -> Generator[Any, None, None]:
        pass
    
    @abc.abstractmethod
    def close(self) -> None:
        pass


class XlsxReadDB(BaseReadDB):

    def __init__(self, filename):
        super().__init__(filename=filename)
        self.reader = load_workbook(filename=filename)

    def read(self) -> Generator[Any, None, None]:
        sheet = self.reader.active
        for row in sheet.values:
            yield row
        else:
            self.close()

    def close(self) -> None:
        self.reader.close()


class BaseOutputWriterDB(metaclass=abc.ABCMeta):

    def __init__(self, filename) -> None:
        self.filename = filename

    @abc.abstractmethod
    def writerow(self, row: Iterable[Any]) -> None:
        pass

    @abc.abstractmethod
    def __enter__(self):
        pass

    @abc.abstractmethod
    def __exit__(self, *args) -> None:
        pass


class XlsxWriterDB(BaseOutputWriterDB):

    def __enter__(self) -> None:
        self.workbook = Workbook(write_only=True)
        self.sheet = self.workbook.create_sheet()
        return self

    def writerow(self, row: Iterable[Any]) -> None:
        self.sheet.append(row)        

    def __exit__(self, *args) -> None:
        self.workbook.save(self.filename)
        self.workbook.close()


class Migrate:

    def __init__(self) -> None:
        self.db1 = None
        self.db2 = None

    def set_db1(self, db1: BaseReadDB) -> None:
        self.db1 = db1

    def set_db2(self, db2: BaseReadDB) -> None:
        self.db2 = db2

    def _iter_products_from_db(self, db: BaseReadDB) -> Generator[Product, None, None]:
        products = db.read()
        for product in products:
            yield Product(description=product[0], price=product[1])

    def run(self, output_db: BaseOutputWriterDB) -> None:
        products_from_db1 = list(self._iter_products_from_db(self.db1))
        products_from_db2 = list(self._iter_products_from_db(self.db2))

        # avoid migrating duplicated
        last_recently_migrated_products = []

        with output_db as output:
            for product_from_db1 in products_from_db1:
                for product_from_db2 in products_from_db2:
                    if product_from_db1 == product_from_db2:

                        if product_from_db1.description in\
                                last_recently_migrated_products:
                            continue

                        output.writerow([
                            product_from_db1.description,
                            product_from_db2.price,
                        ])

                        last_recently_migrated_products.append(
                            product_from_db1.description)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('db1_filename', help='Informe o banco1.csv')
    parser.add_argument('db2_filename', help='Informe o banco2.csv')
    parser.add_argument('-d', '--db_dest',
                        help='Informe um filename do novo banco.')
    parse_args = parser.parse_args()

    migrate = Migrate()
    migrate.set_db1(XlsxReadDB(parse_args.db1_filename))
    migrate.set_db2(XlsxReadDB(parse_args.db2_filename))
    migrate.run(output_db=XlsxWriterDB(parse_args.db_dest))
